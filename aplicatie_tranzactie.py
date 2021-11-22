import  openpyxl as xl
xl.load_workbook('tranzactie.xlsx')
sheet = wb['sheet1']
cell = sheet['a1']
cell = sheet.cell(1,1)
print(sheet.max_row)


for row in range(1 , sheet.max_row + 1):
    cell = sheet.cell(row , 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell( 0 , 4)
    corrected_price_cell.value = corrected_price

wb.save('tranzactie2.xlsx')